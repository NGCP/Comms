import openpyxl
from openpyxl.cell import get_column_letter




def generateMessageData():
    wb = openpyxl.load_workbook('messageTable.xlsx')
    sheetNames = wb.get_sheet_names()
    idLen = 4
    for sheetName in sheetNames:
        sheet = wb.get_sheet_by_name(sheetName)
        maxRow = sheet.max_row
            
        for y  in  range(1,maxRow):    
                    
            cell = 'A' + str(y)
            messageDescription = sheet[cell].value 
            if(str(messageDescription)[0] == '#'):  
                messageID = messageDescription[1:idLen +1]
                messageName = messageDescription[idLen + 2:]
                messageName = messageName.replace(' ', '_')
                nextColCell = 'A' + str(y+1)
                
                y+=2
                if( sheet[nextColCell].value == "Field"):   
                    f.write("\n\t<message name = \"" + messageName + "\" id = \"" + messageID + "\">\n")   
                    
                    while(str(sheet['A' + str(y)].value) != "None"):
                        currentCell = 'C' + str(y)
                        name =  str(sheet[currentCell].value)
                        type = str(sheet['D' + str(y)].value)
                        type = type.replace('_', '') + "_t"
                        f.write("\t\t<field type = \"" + type + "\" name = \"" + name + "\"></field>\n")
                        y +=1                    
                    f.write("\t" +"</message>\n")
               
   

   
#open xml file   
f = open('message_definition.xml', 'w')

#write to xml file with paths and header data
f.write("""<?xml version="1.0"?>
<protocol> 
    <paths>
        <field
            cppPath = "../comnet"
            csharpPath = "../comwrapper"
            nativePath = "native"
        ></field>    
    </paths> 
    
    <!-- to create a pointer or array use uint8_t* and length data -->
    <header>
        <field type = "uint8_t" name = "node_src_id"></field>
        <field type = "uint8_t" name = "node_dest_id"></field>           
        <field type = "uint16_t" name = "message_type"></field> 
        <field type = "uint16_t" name = "message_length"></field>        
    </header>
""")

#read from xlsx sheet and create xml file of message data
generateMessageData()  

#close xml file
f.write("</protocol>") 
f.close()
                
                
            

            